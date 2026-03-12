"""
Fax Download Logger
Saves all successfully downloaded fax file paths to an Excel spreadsheet
"""

import os
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

log = logging.getLogger(__name__)

LOG_DIR = "logs"
LOG_FILE = os.path.join(LOG_DIR, "fax_download_log.xlsx")


class FaxLogger:
    def __init__(self):
        os.makedirs(LOG_DIR, exist_ok=True)

    def save(self, results: list):
        """Save or append results to the Excel audit log"""
        if os.path.exists(LOG_FILE):
            wb = load_workbook(LOG_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Download Log"
            self._write_header(ws)

        for r in results:
            self._write_row(ws, r)

        # Auto-size columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

        wb.save(LOG_FILE)
        log.info(f"Audit log saved to: {LOG_FILE}")

    def _write_header(self, ws):
        headers = ["#", "Timestamp", "Status", "File Path", "Reason"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center")

    def _write_row(self, ws, r: dict):
        row = ws.max_row + 1
        status = r.get("status", "FAILED")
        
        # Green for success, Red for failure
        color = "C6EFCE" if status == "SUCCESS" else "FFC7CE"

        values = [
            r.get("fax_index", ""),
            r.get("timestamp", ""),
            status,
            r.get("filepath", ""),
            r.get("reason", "")
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(wrap_text=True)
